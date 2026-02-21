# Copyright (c) 2026 Mustafa Uzumeri. All rights reserved.

"""
Tabular Data to Markdown Converter

Converts CSV and Excel (.xlsx) files to clean Markdown with tables.

CSV files produce a single Markdown table.  Excel workbooks produce one
section per sheet, each with its own Markdown table.

Usage:
    Single file:
      convert_tabular.py inputs/data.csv
      convert_tabular.py inputs/workbook.xlsx -o outputs/custom.md

    Called from server.py:
      from convert_tabular import convertCsvToMarkdown, convertXlsxToMarkdown
"""

import csv
import re
import time
from pathlib import Path


# ── Helpers ────────────────────────────────────────────────────────────────

def _sanitizeCell(value) -> str:
    """Clean a cell value for safe Markdown table rendering.

    Args:
        value: Raw cell value (any type).

    Returns:
        Sanitized string with pipes escaped and whitespace trimmed.
    """
    if value is None:
        return ""
    text = str(value).strip()
    # Escape pipe characters so they don't break the Markdown table
    text = text.replace("|", "\\|")
    # Collapse internal newlines into spaces
    text = re.sub(r"\s*\n\s*", " ", text)
    return text


def _buildMarkdownTable(headers: list[str], rows: list[list[str]]) -> str:
    """Render headers and rows as a GitHub-flavored Markdown table.

    Args:
        headers: Column header strings.
        rows: List of row data (each row is a list of cell strings).

    Returns:
        Markdown table string.
    """
    if not headers:
        return ""

    safeHeaders = [_sanitizeCell(h) or f"Column {i+1}" for i, h in enumerate(headers)]
    headerLine = "| " + " | ".join(safeHeaders) + " |"
    separatorLine = "| " + " | ".join("---" for _ in safeHeaders) + " |"

    dataLines = []
    for row in rows:
        # Pad row to match header count
        padded = row + [""] * (len(safeHeaders) - len(row))
        safeCells = [_sanitizeCell(c) for c in padded[:len(safeHeaders)]]
        dataLines.append("| " + " | ".join(safeCells) + " |")

    return "\n".join([headerLine, separatorLine] + dataLines)


# ── CSV Converter ──────────────────────────────────────────────────────────

def convertCsvToMarkdown(inputPath: Path, outputPath: Path) -> str:
    """Convert a CSV file to Markdown with a single table.

    Args:
        inputPath: Path to the CSV file.
        outputPath: Path for the output Markdown file.

    Returns:
        The generated Markdown text.
    """
    print("  Mode:     csv (stdlib csv)")

    startTime = time.time()

    # Detect encoding — try UTF-8 first, fall back to latin-1
    for encoding in ("utf-8-sig", "utf-8", "latin-1"):
        try:
            text = inputPath.read_text(encoding=encoding)
            break
        except (UnicodeDecodeError, ValueError):
            continue
    else:
        text = inputPath.read_text(encoding="latin-1", errors="replace")

    # Detect dialect (delimiter, quoting)
    sample = text[:8192]
    try:
        dialect = csv.Sniffer().sniff(sample, delimiters=",;\t|")
    except csv.Error:
        dialect = csv.excel  # default to comma-delimited

    reader = csv.reader(text.splitlines(), dialect)
    allRows = list(reader)

    if not allRows:
        mdText = f"---\nsource_file: {inputPath.name}\n---\n\n# {inputPath.stem}\n\n_Empty CSV file._\n"
    else:
        headers = allRows[0]
        dataRows = allRows[1:]

        table = _buildMarkdownTable(headers, dataRows)

        mdText = (
            f"---\n"
            f"source_file: {inputPath.name}\n"
            f"row_count: {len(dataRows)}\n"
            f"column_count: {len(headers)}\n"
            f"---\n\n"
            f"# {inputPath.stem}\n\n"
            f"{table}\n"
        )

    elapsed = time.time() - startTime

    outputPath.parent.mkdir(parents=True, exist_ok=True)
    outputPath.write_text(mdText, encoding="utf-8")
    print(f"  Done:     {len(mdText):,} chars -> {outputPath} ({elapsed:.1f}s)")

    return mdText


# ── XLSX Converter ─────────────────────────────────────────────────────────

def convertXlsxToMarkdown(inputPath: Path, outputPath: Path) -> str:
    """Convert an Excel workbook to Markdown with one section per sheet.

    Args:
        inputPath: Path to the .xlsx file.
        outputPath: Path for the output Markdown file.

    Returns:
        The generated Markdown text.
    """
    import openpyxl

    print("  Mode:     xlsx (openpyxl)")

    startTime = time.time()

    wb = openpyxl.load_workbook(str(inputPath), read_only=True, data_only=True)

    sections: list[str] = []
    totalRows = 0
    totalSheets = len(wb.sheetnames)

    for sheetName in wb.sheetnames:
        ws = wb[sheetName]
        rows = list(ws.iter_rows(values_only=True))

        if not rows:
            sections.append(f"## {sheetName}\n\n_Empty sheet._\n")
            continue

        # Use the first row as headers
        headers = list(rows[0])
        dataRows = [list(r) for r in rows[1:]]
        totalRows += len(dataRows)

        table = _buildMarkdownTable(headers, dataRows)

        sections.append(
            f"## {sheetName}\n\n"
            f"_{len(dataRows)} rows × {len(headers)} columns_\n\n"
            f"{table}\n"
        )

    wb.close()

    mdText = (
        f"---\n"
        f"source_file: {inputPath.name}\n"
        f"sheet_count: {totalSheets}\n"
        f"total_rows: {totalRows}\n"
        f"---\n\n"
        f"# {inputPath.stem}\n\n"
        + "\n".join(sections)
    )

    elapsed = time.time() - startTime

    outputPath.parent.mkdir(parents=True, exist_ok=True)
    outputPath.write_text(mdText, encoding="utf-8")
    print(f"  Done:     {len(mdText):,} chars -> {outputPath} ({elapsed:.1f}s)")

    return mdText


# ── CLI Entry Point ────────────────────────────────────────────────────────

if __name__ == "__main__":
    import argparse
    import sys

    parser = argparse.ArgumentParser(
        description="Convert CSV or Excel files to Markdown tables",
    )
    parser.add_argument("input", help="Path to a CSV or XLSX file")
    parser.add_argument(
        "-o", "--output",
        help="Output Markdown file path (default: outputs/<stem>.md)",
    )

    args = parser.parse_args()

    inputFile = Path(args.input)
    if not inputFile.exists():
        print(f"Error: File not found: {inputFile}")
        sys.exit(1)

    ext = inputFile.suffix.lower()
    outputDir = Path("outputs")
    outPath = Path(args.output) if args.output else outputDir / f"{inputFile.stem}.md"

    print(f"[1/1] {inputFile.name}")
    if ext == ".csv":
        convertCsvToMarkdown(inputFile, outPath)
    elif ext == ".xlsx":
        convertXlsxToMarkdown(inputFile, outPath)
    else:
        print(f"Error: Unsupported file type: {ext}. Use .csv or .xlsx")
        sys.exit(1)
